from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from soil_vertical_stress_increment.models.methods_enum import MetodosCalculo
from soil_vertical_stress_increment.models.vertical_stress_increment_result import BoussinesqIterationResult, FrolichX2IterationResult, FrolichX4IterationResult, VerticalStressIncrementResults, WestergaardIterationResult
from models.result_value_object import ResultValueObject

class ExportService : 

    def export(self,results:list[ResultValueObject],path:str):
        for result in results:
            print(result.title)
        print(path)
        
        wb = Workbook()
        first_spreadsheet =wb.active
        for result in results:
            spreadsheet:Worksheet=wb.create_sheet(result.title);
            self._dump_data(spreadsheet,result)
        wb.remove(first_spreadsheet)
        wb.save(path)
            
    def _dump_data (self,spreadsheet:Worksheet,result:ResultValueObject):
        spreadsheet.cell(1,1,'Datos de entrada')
        datos_entrada=result.tabla_resultados.input_data.punto_3d_data
        vertices_entrada=result.tabla_resultados.input_data.vertices_data
        resultados  = result.tabla_resultados
        last_row_position = self._draw_input_data(spreadsheet, datos_entrada, vertices_entrada);
        self._dibujar_tabla_resultados(spreadsheet,last_row_position,resultados)
        
    def _dibujar_tabla_resultados(self,spreadsheet:Worksheet, start_row_position:int,resultados:VerticalStressIncrementResults):
        iterations =resultados.get_iteration_results();
        renderer=self._get_row_render_function(resultados.method);
        
        for index , iteration in enumerate(iterations):
            iteration
            
    
    def _draw_input_data(self, spreadsheet:Worksheet, datos_entrada, vertices_entrada):
        spreadsheet.cell(2,1,'q')
        spreadsheet.cell(2,2,datos_entrada.q)
        spreadsheet.cell(3,1,'PUNTO')
        spreadsheet.cell(4,1,'x'); spreadsheet.cell(4,2,datos_entrada.punto.x);
        spreadsheet.cell(5,1,'y');spreadsheet.cell(5,2,datos_entrada.punto.y);
        spreadsheet.cell(6,1,'z');spreadsheet.cell(6,2,datos_entrada.punto.z);
        spreadsheet.cell(7,1,'RIGIDEZ');spreadsheet.cell(7,2,datos_entrada.rigidez[0]);
        spreadsheet.cell(8,1,'VERTICES');
        row_position=0
        for index,vertice in enumerate(vertices_entrada):
            row_position = 8+index
            spreadsheet.cell(row_position,1,vertice.x)
            spreadsheet.cell(row_position,2,vertice.y)
        return row_position
        
    def _get_row_render_function(method:MetodosCalculo):
        def Boussinesq(iteration:BoussinesqIterationResult,row_position:int):
            pass
        def Frolich_x2(iteration:FrolichX2IterationResult,row_position:int):
            pass
        def Frolich_x4(iteration:FrolichX4IterationResult,row_position:int):
            pass
        def Westergaard(iteration:WestergaardIterationResult,row_position:int):
            pass
        
        if   method ==  MetodosCalculo.BOUSSINESQ_X3:
            return Boussinesq
        elif method==MetodosCalculo.FROLICH_X2:
            return Frolich_x2
        elif method==MetodosCalculo.FROLICH_X4:
            return Frolich_x4
        elif method==MetodosCalculo.WESTERGAARD:
            return Westergaard
        