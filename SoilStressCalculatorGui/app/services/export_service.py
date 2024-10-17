from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from soil_vertical_stress_increment.models.methods_enum import MetodosCalculo
from soil_vertical_stress_increment.models.vertical_stress_increment_result import BoussinesqIterationResult, FrolichX2IterationResult, FrolichX4IterationResult, VerticalStressIncrementResults, WestergaardIterationResult
from services.datagrid_parser import StressIncrementResultsToDataGridParser
from models.result_value_object import ResultValueObject

class ExportService : 

    def export(self,results:list[ResultValueObject],path:str):
        for result in results:
            print(result.title)
        print(path)
        
        wb = Workbook()
        first_spreadsheet =wb.active
        for result in results:
            spreadsheet:Worksheet=wb.create_sheet(result.title);
            self._dump_data(spreadsheet,result)
        wb.remove(first_spreadsheet)
        wb.save(path)
            
    def _dump_data (self,spreadsheet:Worksheet,result:ResultValueObject):
        spreadsheet.cell(1,1,'Datos de entrada')
        datos_entrada=result.tabla_resultados.input_data.punto_3d_data
        vertices_entrada=result.tabla_resultados.input_data.vertices_data
        resultados  = result.tabla_resultados
        esfuerzos = result.tabla_esfuerzos
        last_row_position = self._draw_input_data(spreadsheet, datos_entrada, vertices_entrada);
        last_row_position+=2
        last_row_position=self._dibujar_tabla_resultados(spreadsheet,last_row_position,resultados)
        last_row_position+=2
        last_row_position=self._dibujar_tabla_esfuerzos(spreadsheet,last_row_position,esfuerzos)
    
    def _dibujar_tabla_esfuerzos(self,spreadsheet:Worksheet, start_row_position:int,resultados:list[tuple[float, float]]):
        spreadsheet.cell(start_row_position,1,"ESFUERZOS")
        start_row_position+=1;
        spreadsheet.cell(start_row_position,1,"Z")
        spreadsheet.cell(start_row_position,2,"ESFUERZO")
        start_row_position+=1;
        row_pos=0;
        col_pos=0;
        for row_index ,rows in enumerate(resultados):
            for col_index, cell_data in enumerate(rows):
                row_pos=row_index+start_row_position;
                col_pos=col_index+1
                spreadsheet.cell(row_pos,col_pos,cell_data)
        
        return row_pos
        
    def _dibujar_tabla_resultados(self,spreadsheet:Worksheet, start_row_position:int,resultados:VerticalStressIncrementResults):
        iterations =resultados.get_iteration_results();
        esfuerzo = resultados.get_total_result()
        columns=StressIncrementResultsToDataGridParser.getColumns(resultados.method);
        data_grid=StressIncrementResultsToDataGridParser.getDataGrid(resultados.method,iterations)
        matrix= [columns]+data_grid
        row_pos=0
        col_pos=0
        for row_index ,rows in enumerate(matrix):
            for col_index, cell_data in enumerate(rows):
                row_pos=row_index+start_row_position;
                col_pos=col_index+1
                spreadsheet.cell(row_pos,col_pos,cell_data)
        
        row_pos+=1
        lower_right_corner =len(matrix[0])
        spreadsheet.cell(row_pos,lower_right_corner,esfuerzo)
        spreadsheet.cell(row_pos,lower_right_corner-1,'Σ=')
        return row_pos
            
    
    def _draw_input_data(self, spreadsheet:Worksheet, datos_entrada, vertices_entrada):
        spreadsheet.cell(2,1,'q')
        spreadsheet.cell(2,2,datos_entrada.q)
        spreadsheet.cell(3,1,'PUNTO')
        spreadsheet.cell(4,1,'x'); spreadsheet.cell(4,2,datos_entrada.punto.x);
        spreadsheet.cell(5,1,'y');spreadsheet.cell(5,2,datos_entrada.punto.y);
        spreadsheet.cell(6,1,'z');spreadsheet.cell(6,2,datos_entrada.punto.z);
        spreadsheet.cell(7,1,'RIGIDEZ');spreadsheet.cell(7,2,datos_entrada.rigidez[0]);
        spreadsheet.cell(8,1,'VERTICES');
        row_position=0
        for index,vertice in enumerate(vertices_entrada):
            row_position = 8+index
            spreadsheet.cell(row_position,1,vertice.x)
            spreadsheet.cell(row_position,2,vertice.y)
        return row_position
        
   